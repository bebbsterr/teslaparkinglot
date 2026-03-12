"""
=============================================================================
TESLA YARD — PROTOTYPE PARKING ASSIGNMENT SYSTEM  (v3)
=============================================================================
Simulates intelligent vehicle parking for a Tesla production/logistics yard.

HOW TO RUN:
    python3 tesla_yard.py
    python3 tesla_yard.py teslalocations.xlsx

DEPENDENCY:
    pip install openpyxl

YARD LAYOUT:
    7 lanes × 5 spaces = 35 vehicles maximum
    Lanes 1–5  →  Dynamic destination lanes (reassigned based on demand)
    Lane 6     →  Overflow (used when destination lanes are full)
    Lane 7     →  Priority (reserved for high-urgency vehicles)

LANE STRUCTURE (one-way):
    Position 1 = entrance/exit end  →  easiest to retrieve
    Position 5 = deepest in lane    →  blocked by 4 vehicles ahead

=============================================================================
"""

import sys
import os
import random
from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional
from collections import defaultdict

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


# =============================================================================
# CONSTANTS
# =============================================================================

LANE_COUNT      = 7
SPACES_PER_LANE = 5
DYNAMIC_LANES   = 5   # Lanes 1–5 are assigned dynamically to destinations
OVERFLOW_LANE   = 6
PRIORITY_LANE   = 7

MODELS  = ['Model 3', 'Model Y', 'Model S', 'Model X', 'Cybertruck']
COLOURS = ['Pearl White', 'Midnight Silver', 'Deep Blue',
           'Solid Black', 'Ultra Red', 'Stealth Grey', 'Quicksilver']

# Premium models and rare colours get a small config priority boost in scoring
PREMIUM_MODELS = {'Model S', 'Model X', 'Cybertruck'}
RARE_COLOURS   = {'Ultra Red', 'Quicksilver'}

# Scoring weights — applied in ParkingAllocator._score()
# The 8 factors and their relative importance:
W_DESTINATION = 0.30   # Parking in the correct city lane is most important
W_RETRIEVABLE = 0.20   # Front of lane (position 1) is easiest to retrieve
W_URGENCY     = 0.15   # Imminent pickups should be parked most accessibly
W_ORDER_PRIO  = 0.10   # VIP customer orders deserve better placement
W_CONFIG_PRIO = 0.10   # Premium model / rare colour = slight boost
W_CONGESTION  = 0.10   # Prefer lanes that are less congested
W_DISTANCE    = 0.05   # Lane 1 is closest to the loading bay
W_BLOCKING    = 0.10   # Penalty subtracted if lower-urgency cars block exit


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class VehicleOrder:
    """Represents a customer's order for a specific Tesla vehicle."""
    model:               str
    colour:              str
    destination_country: str
    destination_city:    str
    pickup_urgency:      int   # 1–10  (10 = truck arriving today)
    order_priority:      int   # 1–10  (10 = VIP customer)


@dataclass
class Vehicle:
    """A physical vehicle present in the yard."""
    vin:          str
    order:        VehicleOrder
    arrival_time: datetime
    space_id:     Optional[str] = None   # Set once the vehicle is parked
    status:       str = 'incoming'       # 'incoming' | 'parked' | 'retrieved'


@dataclass
class Space:
    """
    One parking space inside a lane.

    Position 1 = nearest the lane entrance (easiest to exit).
    Position 5 = deepest in the lane (blocked by 4 vehicles in front).
    """
    space_id:    str
    lane_number: int
    position:    int
    is_occupied: bool = False
    vehicle_vin: Optional[str] = None

    def is_free(self) -> bool:
        return not self.is_occupied


@dataclass
class Lane:
    """
    A one-way parking lane holding up to SPACES_PER_LANE vehicles.

    lane_type   : 'dynamic' | 'overflow' | 'priority'
    assigned_to : destination key, e.g. 'Paris, France'  (dynamic lanes only)
    """
    lane_number:  int
    lane_type:    str
    assigned_to:  str = ''
    spaces:       list = field(default_factory=list)

    def available_spaces(self) -> list:
        return [s for s in self.spaces if s.is_free()]

    def occupied_spaces(self) -> list:
        return [s for s in self.spaces if s.is_occupied]

    def occupancy_rate(self) -> float:
        return len(self.occupied_spaces()) / len(self.spaces)

    def label(self) -> str:
        """Short human-readable label for this lane."""
        if self.lane_type == 'overflow':
            return f"Lane {self.lane_number} [OVERFLOW]"
        if self.lane_type == 'priority':
            return f"Lane {self.lane_number} [PRIORITY]"
        dest = self.assigned_to or '(unassigned)'
        return f"Lane {self.lane_number} [{dest}]"


# =============================================================================
# YARD
# =============================================================================

class Yard:
    """
    The full parking yard: 7 lanes × 5 spaces = 35 total spaces.

    Attributes
    ----------
    lanes    : dict of lane_number → Lane
    vehicles : dict of VIN → Vehicle  (all vehicles ever seen, any status)
    """

    def __init__(self):
        self.lanes:    dict[int, Lane]    = {}
        self.vehicles: dict[str, Vehicle] = {}
        self._build_lanes()

    def _build_lanes(self):
        for n in range(1, LANE_COUNT + 1):
            if n == PRIORITY_LANE:
                lane_type = 'priority'
            elif n == OVERFLOW_LANE:
                lane_type = 'overflow'
            else:
                lane_type = 'dynamic'

            lane = Lane(lane_number=n, lane_type=lane_type)
            for pos in range(1, SPACES_PER_LANE + 1):
                lane.spaces.append(Space(
                    space_id=f"L{n}-P{pos}",
                    lane_number=n,
                    position=pos
                ))
            self.lanes[n] = lane

    # Convenience accessors ──────────────────────────────────────────────────

    def dynamic_lanes(self) -> list:
        return [self.lanes[n] for n in range(1, DYNAMIC_LANES + 1)]

    def overflow_lane(self) -> Lane:
        return self.lanes[OVERFLOW_LANE]

    def priority_lane(self) -> Lane:
        return self.lanes[PRIORITY_LANE]

    def get_space(self, space_id: str) -> Optional[Space]:
        for lane in self.lanes.values():
            for s in lane.spaces:
                if s.space_id == space_id:
                    return s
        return None

    def total_occupied(self) -> int:
        return sum(1 for lane in self.lanes.values() for s in lane.spaces if s.is_occupied)

    def total_capacity(self) -> int:
        return LANE_COUNT * SPACES_PER_LANE

    def is_full(self) -> bool:
        return self.total_occupied() >= self.total_capacity()

    def parked_vehicles(self) -> list:
        return [v for v in self.vehicles.values() if v.status == 'parked']


# =============================================================================
# LANE ALLOCATOR
# =============================================================================

class LaneAllocator:
    """
    Dynamically assigns the 5 destination lanes to cities based on demand.

    Demand = number of vehicles currently parked for each destination.
    Destinations with more vehicles get more lanes.

    Example
    -------
    Parked:  Paris×9, London×7, Manchester×4, Berlin×3, Amsterdam×2
    Result:  Lane1→Paris, Lane2→Paris, Lane3→London,
             Lane4→London, Lane5→Manchester
    """

    def recalculate(self, yard: Yard) -> dict:
        """
        Recompute lane assignments from current yard state.
        Returns a dict of {lane_number: destination_key}.
        """
        # Count how many vehicles are parked per destination
        demand: dict[str, int] = defaultdict(int)
        for v in yard.parked_vehicles():
            key = f"{v.order.destination_city}, {v.order.destination_country}"
            demand[key] += 1

        if not demand:
            for lane in yard.dynamic_lanes():
                lane.assigned_to = ''
            return {}

        # Sort destinations by demand, highest first
        ranked = sorted(demand.items(), key=lambda x: x[1], reverse=True)
        total  = sum(demand.values())

        # Distribute lanes proportionally (each destination gets at least 1)
        allocation = self._proportional_split(ranked, total)

        # Apply to lanes 1–5
        assignment = {}
        lane_num   = 1
        for dest, lanes_count in allocation:
            for _ in range(lanes_count):
                if lane_num > DYNAMIC_LANES:
                    break
                yard.lanes[lane_num].assigned_to = dest
                assignment[lane_num] = dest
                lane_num += 1
            if lane_num > DYNAMIC_LANES:
                break

        # Any lanes with no assignment (fewer destinations than lanes)
        for n in range(lane_num, DYNAMIC_LANES + 1):
            yard.lanes[n].assigned_to = ''

        return assignment

    def _proportional_split(self, ranked: list, total: int) -> list:
        """
        Given a ranked list of (destination, count), return a list of
        (destination, lanes_allocated) using proportional rounding.
        Total allocated lanes always equals DYNAMIC_LANES.
        """
        n_dests    = min(len(ranked), DYNAMIC_LANES)
        allocation = []
        remaining  = DYNAMIC_LANES

        for i, (dest, count) in enumerate(ranked[:n_dests]):
            dests_left = n_dests - i
            # Each destination gets at least 1 lane; scale by share of total
            share = max(1, round((count / total) * DYNAMIC_LANES))
            share = min(share, remaining - (dests_left - 1))
            allocation.append((dest, share))
            remaining -= share

        return allocation


# =============================================================================
# PARKING ALLOCATOR
# =============================================================================

class ParkingAllocator:
    """
    Finds the best available parking space for an incoming vehicle.

    Uses a weighted scoring system across 8 factors:
      destination match, retrievability, urgency, order priority,
      config priority, lane congestion, distance, blocking risk.
    """

    def assign(self, vehicle: Vehicle, yard: Yard) -> tuple:
        """
        Score every candidate space and assign the best one.

        Returns (space, lane, score, breakdown) on success,
                (None, None, 0, {}) if the yard is full.
        """
        best = (-1.0, None, None, {})

        for lane, space in self._candidates(vehicle, yard):
            score, breakdown = self._score(vehicle, space, lane, yard)
            if score > best[0]:
                best = (score, space, lane, breakdown)

        score, space, lane, breakdown = best
        if space is None:
            return None, None, 0, {}

        # Commit the assignment
        space.is_occupied  = True
        space.vehicle_vin  = vehicle.vin
        vehicle.space_id   = space.space_id
        vehicle.status     = 'parked'
        yard.vehicles[vehicle.vin] = vehicle

        return space, lane, score, breakdown

    # ── Candidate generation ─────────────────────────────────────────────────

    def _candidates(self, vehicle: Vehicle, yard: Yard) -> list:
        """
        Return (lane, space) pairs to evaluate, in preference order:
        1. Priority lane  — if urgency ≥ 8
        2. Matching destination lane(s)
        3. Overflow lane
        4. Any dynamic lane (fallback if no lane is assigned yet)
        """
        dest_key   = f"{vehicle.order.destination_city}, {vehicle.order.destination_country}"
        candidates = []

        if vehicle.order.pickup_urgency >= 8:
            for sp in yard.priority_lane().available_spaces():
                candidates.append((yard.priority_lane(), sp))

        for lane in yard.dynamic_lanes():
            if lane.assigned_to == dest_key:
                for sp in lane.available_spaces():
                    candidates.append((lane, sp))

        for sp in yard.overflow_lane().available_spaces():
            candidates.append((yard.overflow_lane(), sp))

        # If no lane is assigned yet for this destination, try all dynamic lanes
        if not candidates:
            for lane in yard.dynamic_lanes():
                for sp in lane.available_spaces():
                    candidates.append((lane, sp))

        return candidates

    # ── Scoring ──────────────────────────────────────────────────────────────

    def _score(self, vehicle: Vehicle, space: Space, lane: Lane, yard: Yard) -> tuple:
        """
        Compute the weighted parking score for one (vehicle, space, lane).
        Returns (total_score, breakdown_dict).
        """
        order    = vehicle.order
        dest_key = f"{order.destination_city}, {order.destination_country}"

        # 1. Destination match
        if lane.lane_type in ('overflow', 'priority'):
            dest_score = 0.2                      # Fallback zones score low
        elif lane.assigned_to == dest_key:
            dest_score = 1.0                      # Perfect match
        elif order.destination_country in lane.assigned_to:
            dest_score = 0.4                      # Same country, different city
        else:
            dest_score = 0.0

        # 2. Retrievability — position 1 (entrance) is easiest to get out
        ret_score = (SPACES_PER_LANE - space.position + 1) / SPACES_PER_LANE

        # 3. Pickup urgency
        urgency_score = order.pickup_urgency / 10.0

        # 4. Order priority
        priority_score = order.order_priority / 10.0

        # 5. Configuration priority (model type + colour rarity)
        config_score = self._config_score(order)

        # 6. Congestion — prefer less-full lanes
        congestion_score = 1.0 - lane.occupancy_rate()

        # 7. Distance — Lane 1 is closest to the loading bay
        distance_score = (LANE_COUNT - lane.lane_number + 1) / LANE_COUNT

        # 8. Blocking risk — penalty for low-urgency cars between this space and exit
        blocking_penalty = self._blocking_risk(vehicle, space, lane, yard)

        total = (
            W_DESTINATION * dest_score
          + W_RETRIEVABLE * ret_score
          + W_URGENCY     * urgency_score
          + W_ORDER_PRIO  * priority_score
          + W_CONFIG_PRIO * config_score
          + W_CONGESTION  * congestion_score
          + W_DISTANCE    * distance_score
          - W_BLOCKING    * blocking_penalty
        )
        total = round(max(0.0, min(1.0, total)), 4)

        breakdown = {
            'destination': (W_DESTINATION, dest_score),
            'retrievable': (W_RETRIEVABLE, ret_score),
            'urgency':     (W_URGENCY,     urgency_score),
            'order_prio':  (W_ORDER_PRIO,  priority_score),
            'config_prio': (W_CONFIG_PRIO, config_score),
            'congestion':  (W_CONGESTION,  congestion_score),
            'distance':    (W_DISTANCE,    distance_score),
            'blocking':    (W_BLOCKING,    blocking_penalty),
        }
        return total, breakdown

    def _config_score(self, order: VehicleOrder) -> float:
        """
        Score based on how premium this vehicle configuration is.

        Baseline:         0.4
        Premium model:   +0.4  (Model S / X / Cybertruck)
        Rare colour:     +0.2  (Ultra Red / Quicksilver)
        Max:              1.0
        """
        score = 0.4
        if order.model in PREMIUM_MODELS:
            score += 0.4
        if order.colour in RARE_COLOURS:
            score += 0.2
        return min(1.0, score)

    def _blocking_risk(self, vehicle: Vehicle, space: Space,
                       lane: Lane, yard: Yard) -> float:
        """
        Estimate the risk that this vehicle will be hard to retrieve.

        In a one-way lane, all vehicles at positions 1 to (position-1)
        must exit before this vehicle can leave.
        
        The penalty is higher when those blocking vehicles have LOWER urgency
        than this vehicle — meaning they won't naturally leave first.

        Returns a value between 0.0 (no risk) and 1.0 (fully blocked).
        """
        if space.position == 1:
            return 0.0   # Nothing blocks position 1

        blockers = [s for s in lane.occupied_spaces() if s.position < space.position]
        if not blockers:
            return 0.0

        low_urgency_count = 0
        for s in blockers:
            blocking_vehicle = yard.vehicles.get(s.vehicle_vin)
            if blocking_vehicle:
                if blocking_vehicle.order.pickup_urgency < vehicle.order.pickup_urgency:
                    low_urgency_count += 1

        return low_urgency_count / len(blockers)


# =============================================================================
# TRUCK RETRIEVAL SIMULATOR
# =============================================================================

class TruckRetrieval:
    """
    Simulates a truck arriving to collect vehicles for a given destination.

    Works out which vehicles can be retrieved directly and which require
    other vehicles to be temporarily moved out of the way first.
    """

    def simulate(self, country: str, city: str, count: int, yard: Yard) -> dict:
        """
        Plan a retrieval of up to `count` vehicles for (country, city).

        Returns
        -------
        dict with:
          'requested'  : how many were requested
          'found'      : how many matching vehicles exist
          'sequence'   : list of step dicts (one per vehicle to retrieve)
          'temp_moves' : vehicles that must temporarily move
        """
        matching = [
            v for v in yard.parked_vehicles()
            if v.order.destination_city    == city
            and v.order.destination_country == country
        ]

        if not matching:
            return {'requested': count, 'found': 0, 'sequence': [], 'temp_moves': []}

        # Retrieve highest-urgency first; within same urgency, nearest exit first
        matching.sort(key=lambda v: (
            -v.order.pickup_urgency,
            yard.get_space(v.space_id).position if v.space_id else 99
        ))

        to_retrieve = matching[:count]
        sequence    = []
        temp_moves  = []

        for vehicle in to_retrieve:
            space = yard.get_space(vehicle.space_id)
            if space is None:
                continue

            lane = yard.lanes[space.lane_number]

            # Blockers: occupied spaces between this vehicle and the lane exit
            # that are NOT themselves being retrieved in this batch
            blockers = []
            for s in lane.occupied_spaces():
                if s.position < space.position:
                    blocking_v = yard.vehicles.get(s.vehicle_vin)
                    if blocking_v and blocking_v not in to_retrieve:
                        blockers.append((s.position, blocking_v, s))

            blockers.sort(key=lambda x: x[0])

            sequence.append({
                'vehicle':  vehicle,
                'space':    space,
                'lane':     lane,
                'blockers': blockers,
            })
            temp_moves.extend(bv for _, bv, _ in blockers)

        return {
            'requested':  count,
            'found':      len(to_retrieve),
            'sequence':   sequence,
            'temp_moves': list(set(temp_moves)),
        }

    def execute(self, result: dict, yard: Yard):
        """
        Commit a planned retrieval: free the spaces and mark vehicles retrieved.
        Only call this after displaying the result and confirming with the user.
        """
        for step in result['sequence']:
            vehicle = step['vehicle']
            space   = step['space']
            space.is_occupied = False
            space.vehicle_vin = None
            vehicle.space_id  = None
            vehicle.status    = 'retrieved'


# =============================================================================
# EXCEL / DESTINATION LOADER
# =============================================================================

def load_destinations(filepath: str) -> list:
    """
    Load country/city pairs from an Excel file (teslalocations.xlsx).

    Looks for a sheet named 'Destination_Hubs' (or the first sheet).
    Expects columns: 'Country' and 'City' (or 'City / Hub').

    Falls back to a built-in European default list if the file is missing
    or cannot be read.
    """
    if not EXCEL_AVAILABLE:
        print("  ⚠  openpyxl not installed — using built-in destinations.")
        print("     Run:  pip install openpyxl")
        return _default_destinations()

    if not os.path.exists(filepath):
        print(f"  ⚠  File not found: '{filepath}' — using built-in destinations.")
        return _default_destinations()

    try:
        wb = openpyxl.load_workbook(filepath)

        # Try known sheet names, fall back to first sheet
        sheet_name = next(
            (name for name in ['Destination_Hubs', 'Destinations', 'Sheet1']
             if name in wb.sheetnames),
            wb.sheetnames[0]
        )
        ws      = wb[sheet_name]
        rows    = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip().lower() if h else '' for h in rows[0]]

        col_country = next((i for i, h in enumerate(headers) if 'country' in h), None)
        col_city    = next((i for i, h in enumerate(headers) if 'city' in h or 'hub' in h), None)

        if col_country is None or col_city is None:
            print("  ⚠  Could not identify Country/City columns — using built-in destinations.")
            return _default_destinations()

        destinations = []
        seen = set()
        for row in rows[1:]:
            if row[col_country] and row[col_city]:
                c  = str(row[col_country]).strip()
                ci = str(row[col_city]).strip()
                if c.lower() != 'country' and (c, ci) not in seen:
                    seen.add((c, ci))
                    destinations.append({'country': c, 'city': ci})

        if destinations:
            print(f"  ✔  Loaded {len(destinations)} destinations from '{filepath}'")
            return destinations

        print("  ⚠  No destinations found in file — using built-in destinations.")
        return _default_destinations()

    except Exception as exc:
        print(f"  ⚠  Could not read '{filepath}': {exc}")
        print("     Using built-in destinations.")
        return _default_destinations()


def _default_destinations() -> list:
    """Built-in European city fallback list."""
    return [
        {'country': 'France',      'city': 'Paris'},
        {'country': 'France',      'city': 'Lyon'},
        {'country': 'UK',          'city': 'London'},
        {'country': 'UK',          'city': 'Manchester'},
        {'country': 'UK',          'city': 'Birmingham'},
        {'country': 'Germany',     'city': 'Berlin'},
        {'country': 'Germany',     'city': 'Munich'},
        {'country': 'Netherlands', 'city': 'Amsterdam'},
        {'country': 'Netherlands', 'city': 'Rotterdam'},
        {'country': 'Spain',       'city': 'Madrid'},
        {'country': 'Spain',       'city': 'Barcelona'},
        {'country': 'Italy',       'city': 'Rome'},
        {'country': 'Italy',       'city': 'Milan'},
    ]


# =============================================================================
# HELPER / INPUT UTILITIES
# =============================================================================

def generate_vin() -> str:
    chars = "ABCDEFGHJKLMNPRSTUVWXYZ0123456789"
    return "5YJ" + ''.join(random.choices(chars, k=14))


def urgency_label(u: int) -> str:
    if u >= 9: return "Truck arriving today / imminent"
    if u >= 7: return "Truck arriving within 1–2 days"
    if u >= 5: return "Truck arriving this week"
    if u >= 3: return "Truck arriving within 2 weeks"
    return "No firm collection date"


def pick_from_list(prompt: str, options: list, display=None) -> object:
    """Show a numbered list and return the chosen item."""
    print(f"\n{prompt}")
    for i, opt in enumerate(options, 1):
        label = display(opt) if display else str(opt)
        print(f"  {i:>2}.  {label}")
    while True:
        raw = input("\n  Choice: ").strip()
        if raw.lower() == 'quit':
            sys.exit(0)
        if raw.isdigit() and 1 <= int(raw) <= len(options):
            return options[int(raw) - 1]
        print(f"  Please enter a number between 1 and {len(options)}.")


def pick_int(prompt: str, lo: int, hi: int) -> int:
    """Prompt for an integer in [lo, hi]."""
    while True:
        raw = input(f"  {prompt} ({lo}–{hi}): ").strip()
        if raw.lower() == 'quit':
            sys.exit(0)
        if raw.isdigit() and lo <= int(raw) <= hi:
            return int(raw)
        print(f"  Please enter a whole number between {lo} and {hi}.")


# =============================================================================
# DISPLAY FUNCTIONS
# =============================================================================

def display_yard(yard: Yard):
    """Print a visual map of all 7 lanes and their spaces."""
    print("\n" + "═"*66)
    print("  YARD OVERVIEW")
    print("═"*66)
    print(f"  {yard.total_occupied()} / {yard.total_capacity()} spaces occupied\n")

    for n in range(1, LANE_COUNT + 1):
        lane = yard.lanes[n]
        bar  = "█" * len(lane.occupied_spaces()) + "░" * len(lane.available_spaces())
        print(f"  {lane.label():<32}  {bar}  ({lane.occupancy_rate()*100:.0f}%)")
        for space in lane.spaces:
            if space.is_occupied:
                v = yard.vehicles.get(space.vehicle_vin)
                if v:
                    info = f"{v.order.model:<12}  {v.order.colour:<16}  → {v.order.destination_city}"
                    print(f"    P{space.position}  {space.space_id:<8}  {info}  [{v.vin[-6:]}]")
                else:
                    print(f"    P{space.position}  {space.space_id:<8}  [occupied — vehicle data missing]")
            else:
                print(f"    P{space.position}  {space.space_id:<8}  ·  (empty)")
        print()


def display_lane_allocation(yard: Yard):
    """Show the current dynamic lane assignments."""
    print("─"*50)
    print("  LANE ALLOCATION")
    print("─"*50)
    for n in range(1, DYNAMIC_LANES + 1):
        lane = yard.lanes[n]
        dest = lane.assigned_to if lane.assigned_to else "(unassigned)"
        occ  = len(lane.occupied_spaces())
        print(f"  Lane {n}  →  {dest:<25}  {occ}/{SPACES_PER_LANE} occupied")
    occ6 = len(yard.overflow_lane().occupied_spaces())
    occ7 = len(yard.priority_lane().occupied_spaces())
    print(f"  Lane {OVERFLOW_LANE}  →  OVERFLOW                   {occ6}/{SPACES_PER_LANE} occupied")
    print(f"  Lane {PRIORITY_LANE}  →  PRIORITY                   {occ7}/{SPACES_PER_LANE} occupied")
    print()


def display_vehicle_table(yard: Yard):
    """Print a table of all currently parked vehicles."""
    parked = yard.parked_vehicles()
    if not parked:
        print("\n  No vehicles currently parked.\n")
        return

    print("\n" + "─"*92)
    print(f"  {'VIN':<17}  {'Model':<12}  {'Colour':<16}  {'Destination':<24}  "
          f"{'Urg':>3}  {'Pri':>3}  {'Space'}")
    print("─"*92)
    for v in sorted(parked, key=lambda x: x.space_id or ''):
        dest = f"{v.order.destination_city}, {v.order.destination_country}"
        print(f"  {v.vin:<17}  {v.order.model:<12}  {v.order.colour:<16}  "
              f"{dest:<24}  {v.order.pickup_urgency:>3}  "
              f"{v.order.order_priority:>3}  {v.space_id}")
    print()


def display_assignment_result(vehicle: Vehicle, space: Space, lane: Lane,
                               score: float, breakdown: dict):
    """Explain why a vehicle was assigned to a particular space."""
    print("\n" + "═"*62)
    print("  PARKING ASSIGNMENT RESULT")
    print("═"*62)

    dest = f"{vehicle.order.destination_city}, {vehicle.order.destination_country}"

    print(f"\n  Vehicle     :  {vehicle.order.model}  |  {vehicle.order.colour}")
    print(f"  VIN         :  {vehicle.vin}")
    print(f"  Destination :  {dest}")
    print(f"  Urgency     :  {vehicle.order.pickup_urgency}/10 — {urgency_label(vehicle.order.pickup_urgency)}")

    if space.position == 1:
        pos_note = "Front of lane — exits freely, nothing blocking"
    else:
        pos_note = f"{space.position - 1} vehicle(s) between here and lane exit"

    print(f"\n  ┌──────────────────────────────────────────────────┐")
    print(f"  │  Space assigned  :  {space.space_id:<30}│")
    print(f"  │  Lane            :  {lane.label()[:30]:<30}│")
    print(f"  │  Position        :  {space.position} / {SPACES_PER_LANE}                          │")
    print(f"  │  {pos_note:<50}│")
    print(f"  │  Parking score   :  {score:.4f} / 1.0000                  │")
    print(f"  └──────────────────────────────────────────────────┘")

    # Score breakdown table
    print("\n  Score breakdown:")
    print(f"  {'Factor':<22}  {'Weight':>6}  {'Raw':>6}  {'Contribution':>12}")
    print(f"  {'─'*52}")
    factors = [
        ('Destination match',  'destination'),
        ('Retrievability',     'retrievable'),
        ('Pickup urgency',     'urgency'),
        ('Order priority',     'order_prio'),
        ('Config priority',    'config_prio'),
        ('Lane congestion',    'congestion'),
        ('Distance',           'distance'),
    ]
    for label, key in factors:
        w, raw = breakdown.get(key, (0, 0))
        print(f"  {label:<22}  {w:>6.2f}  {raw:>6.2f}  {w * raw:>12.4f}")

    w_b, raw_b = breakdown.get('blocking', (0, 0))
    print(f"  {'Blocking penalty':<22}  {w_b:>6.2f}  {raw_b:>6.2f}  {-w_b * raw_b:>12.4f}")
    print(f"  {'─'*52}")
    print(f"  {'TOTAL':>38}  {score:>12.4f}")
    print()


def display_retrieval_result(result: dict, yard: Yard):
    """Print the planned retrieval sequence for a truck arrival."""
    print("\n" + "═"*62)
    print("  TRUCK RETRIEVAL SIMULATION")
    print("═"*62)
    print(f"\n  Requested : {result['requested']}")
    print(f"  Found     : {result['found']}")

    if result['found'] == 0:
        print("\n  No matching vehicles found for that destination.\n")
        return

    if result['temp_moves']:
        print(f"\n  ⚠  {len(result['temp_moves'])} vehicle(s) must temporarily move:")
        for v in result['temp_moves']:
            s = yard.get_space(v.space_id)
            sp_id = s.space_id if s else '?'
            print(f"     • {v.vin[-8:]}  {v.order.model:<12}  "
                  f"space {sp_id}  → dest: {v.order.destination_city} "
                  f"(urgency {v.order.pickup_urgency})")
    else:
        print("\n  ✔  No temporary moves needed — all vehicles exit freely.")

    print(f"\n  Retrieval sequence:")
    for i, step in enumerate(result['sequence'], 1):
        v        = step['vehicle']
        sp       = step['space']
        lane     = step['lane']
        blockers = step['blockers']

        print(f"\n  Step {i}:")
        print(f"    Vehicle  :  {v.vin[-8:]}  {v.order.model}  {v.order.colour}")
        print(f"    Space    :  {sp.space_id}  in {lane.label()}")
        print(f"    Urgency  :  {v.order.pickup_urgency}/10 — {urgency_label(v.order.pickup_urgency)}")

        if blockers:
            print(f"    ⚠  Blocked — {len(blockers)} vehicle(s) must clear first:")
            for pos, bv, bs in blockers:
                print(f"       → {bv.vin[-8:]}  at {bs.space_id}  "
                      f"({bv.order.destination_city}, urgency {bv.order.pickup_urgency})")
        else:
            print(f"    ✔  Clear — no vehicles blocking exit")

    print()


# =============================================================================
# MAIN INTERACTIVE LOOP
# =============================================================================

def run(spreadsheet_path: str):
    print("\n" + "█"*62)
    print("  TESLA YARD — PROTOTYPE PARKING SYSTEM  (v3)")
    print("█"*62 + "\n")

    # Initialise all components
    destinations    = load_destinations(spreadsheet_path)
    yard            = Yard()
    allocator       = ParkingAllocator()
    lane_allocator  = LaneAllocator()
    retrieval       = TruckRetrieval()

    countries         = sorted(set(d['country'] for d in destinations))
    cities_by_country = defaultdict(list)
    for d in destinations:
        cities_by_country[d['country']].append(d['city'])

    print(f"\n  Yard ready — {yard.total_capacity()} spaces  |  "
          f"{len(destinations)} destinations loaded")
    print("  Type 'quit' at any prompt to exit.\n")

    MENU = [
        "Add vehicle to yard",
        "View yard layout",
        "View vehicle list",
        "Simulate truck arrival",
        "Recalculate lane allocation",
        "Quit",
    ]

    while True:
        print("\n" + "─"*62)
        print("  MAIN MENU")
        print("─"*62)
        for i, item in enumerate(MENU, 1):
            print(f"  {i}.  {item}")
        print()

        choice = pick_int("Choose", 1, len(MENU))

        # ── 1. Add vehicle ────────────────────────────────────────────────
        if choice == 1:
            if yard.is_full():
                print("\n  ⚠  Yard is full — cannot add more vehicles.")
                continue

            print("\n── NEW VEHICLE ─────────────────────────────────────────")

            vin_raw = input("  VIN (Enter to auto-generate): ").strip()
            if vin_raw.lower() == 'quit':
                break
            vin = vin_raw.upper() if vin_raw else generate_vin()
            print(f"  VIN: {vin}")

            if vin in yard.vehicles:
                print("  ⚠  A vehicle with that VIN is already in the yard.")
                continue

            model   = pick_from_list("  Model:", MODELS)
            colour  = pick_from_list("  Colour:", COLOURS)
            country = pick_from_list("  Destination country:", countries)
            city    = pick_from_list(
                f"  Destination city ({country}):",
                sorted(cities_by_country[country])
            )

            print("\n  Pickup urgency:  1 = no date, 10 = truck arriving today")
            urgency  = pick_int("  Urgency", 1, 10)
            priority = pick_int("  Order priority  (1 = standard, 10 = VIP)", 1, 10)

            order = VehicleOrder(
                model=model, colour=colour,
                destination_country=country, destination_city=city,
                pickup_urgency=urgency, order_priority=priority,
            )
            vehicle = Vehicle(vin=vin, order=order, arrival_time=datetime.now())

            space, lane, score, breakdown = allocator.assign(vehicle, yard)

            if space is None:
                print("\n  ⚠  No suitable space found.")
            else:
                lane_allocator.recalculate(yard)
                display_assignment_result(vehicle, space, lane, score, breakdown)

        # ── 2. View yard layout ───────────────────────────────────────────
        elif choice == 2:
            display_yard(yard)
            display_lane_allocation(yard)

        # ── 3. Vehicle list ───────────────────────────────────────────────
        elif choice == 3:
            display_vehicle_table(yard)

        # ── 4. Truck arrival simulation ───────────────────────────────────
        elif choice == 4:
            print("\n── TRUCK ARRIVAL ───────────────────────────────────────")
            country = pick_from_list("  Collecting for country:", countries)
            city    = pick_from_list(
                f"  City ({country}):",
                sorted(cities_by_country[country])
            )
            max_collect = max(1, len([
                v for v in yard.parked_vehicles()
                if v.order.destination_city == city
                and v.order.destination_country == country
            ]))
            count = pick_int("  How many vehicles to collect", 1, max(1, max_collect))

            result = retrieval.simulate(country, city, count, yard)
            display_retrieval_result(result, yard)

            if result['found'] > 0:
                confirm = input("  Execute retrieval? (yes / no): ").strip().lower()
                if confirm in ('yes', 'y'):
                    retrieval.execute(result, yard)
                    lane_allocator.recalculate(yard)
                    print(f"\n  ✔  {result['found']} vehicle(s) marked as retrieved.")

        # ── 5. Recalculate lane allocation ────────────────────────────────
        elif choice == 5:
            lane_allocator.recalculate(yard)
            print("\n  Lane allocation recalculated.")
            display_lane_allocation(yard)

        # ── 6. Quit ───────────────────────────────────────────────────────
        elif choice == 6:
            break

    print("\n  Goodbye!\n")


# =============================================================================
# ENTRY POINT
# =============================================================================

if __name__ == "__main__":
    default_path = "teslalocations.xlsx"
    path = sys.argv[1] if len(sys.argv) > 1 else default_path

    # Also check the uploads directory (for development/demo use)
    if not os.path.exists(path):
        alt = f"/mnt/user-data/uploads/{os.path.basename(path)}"
        if os.path.exists(alt):
            path = alt

    run(path)
